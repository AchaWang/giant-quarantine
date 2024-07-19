import streamlit as st
import openpyxl
import string
import pyperclip

def excel_columns():
    return list(string.ascii_uppercase) + [i+j for i in string.ascii_uppercase for j in string.ascii_uppercase]

def copy_to_clipboard(text):
    pyperclip.copy(text)

def main():
    st.title('Excel 檔案 Column合併')

    # 檔案上傳
    uploaded_file = st.file_uploader("請選擇一個Excel檔案", type=["xlsx", "xls"])

    if uploaded_file is not None:
        # 讀取Excel檔案
        wb = openpyxl.load_workbook(uploaded_file)
        
        # 選擇工作表
        sheet_name = st.selectbox('請選擇工作表', wb.sheetnames)
        
        # 讀取選定的工作表
        ws = wb[sheet_name]
        
        # 獲取工作表的最大列數
        max_column = ws.max_column
        
        # 創建欄位選項列表
        columns = excel_columns()[:max_column]
        
        # 選擇列
        column = st.selectbox('請選擇欄位(column)', columns)
        
        # 獲取選定列的所有值，忽略第一筆數據
        values = [cell.value for cell in list(ws[column])[2:] if cell.value is not None]
        
        # 顯示選定的列數據（忽略第一筆）
        st.write(f'您選擇的欄位(column) "{column}" 的數據（忽略"Frame_no"）:')
        st.write(values)

        # 合併列內容並以逗號隔開，忽略第一筆數據
        merged_content = ','.join(map(str, values))
        
        # 顯示合併後的內容
        st.write('合併後的內容（以逗號隔開，忽略第一筆數據）:')
        st.text_area("合併結果", merged_content, height=200)

        # 添加複製按鈕
        #if st.button('複製合併結果'):
        #    copy_to_clipboard(merged_content)
        #    st.success('已複製到剪貼簿！')

if __name__ == '__main__':
    main()
